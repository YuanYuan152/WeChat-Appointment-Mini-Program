"""Web 管理端来访、咨询师和订单 Excel 导入导出服务。"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from models import (
    AppAccount,
    AppConsultation,
    AppCounselorProfile,
    AppOrder,
    AppRoleBinding,
    AppSchedule,
)
from pricing_service import default_base_price_cents_for_type
from counselor_avatar import DEFAULT_COUNSELOR_PUBLIC_AVATAR
from schedule_meta import parse_center_id, schedule_note
from staff_remark_service import get_staff_remarks_map, set_staff_remark
from user_role_meta import (
    COUNSELOR_TYPES,
    PATIENT_SOURCES,
    PATIENT_SOURCE_DETAILS,
    normalize_patient_source,
)


PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
ORDER_STATUSES = (
    "已预约但未咨询",
    "已咨询未填写咨询记录",
    "已退款",
    "已取消未退款",
)
CONSULTATION_MODES = ("线下", "视频")
LOCATIONS = ("杨浦咨询中心", "浦东咨询中心", "线上")
SIGN_STATUSES = ("已签约", "未签约")
GENDERS = ("男", "女")

SOURCE_BY_LABEL = {label: code for code, label in PATIENT_SOURCES.items()}
COUNSELOR_TYPE_BY_LABEL = {label: code for code, label in COUNSELOR_TYPES.items()}
LOCATION_TO_CENTER = {
    "杨浦咨询中心": "yangpu",
    "浦东咨询中心": "pudong",
    "线上": "video",
}
CENTER_TO_LOCATION = {value: key for key, value in LOCATION_TO_CENTER.items()}


@dataclass(frozen=True)
class ColumnDef:
    header: str
    key: str
    required: bool = False
    choices: tuple[str, ...] = ()


VISITOR_COLUMNS = (
    ColumnDef("来访手机号【必填】", "mobile", True),
    ColumnDef("来访姓名", "real_name"),
    ColumnDef("来访昵称", "nickname"),
    ColumnDef("签约状态", "contract_status", choices=SIGN_STATUSES),
    ColumnDef("绑定咨询师的名字", "bound_counselor_name"),
    ColumnDef("绑定的咨询师电话", "bound_counselor_mobile"),
    ColumnDef("来访类别【必填】", "patient_source", True, tuple(PATIENT_SOURCES.values())),
    ColumnDef("来访来源", "patient_source_detail", choices=PATIENT_SOURCE_DETAILS),
    ColumnDef("来访备注", "remark"),
)

COUNSELOR_COLUMNS = (
    ColumnDef("咨询师手机号【必填】", "mobile", True),
    ColumnDef("咨询师姓名【必填】", "name", True),
    ColumnDef("咨询师性别", "gender", choices=GENDERS),
    ColumnDef("咨询师类别", "counselor_type", choices=tuple(COUNSELOR_TYPES.values())),
    ColumnDef("咨询师基础价格", "billing"),
    ColumnDef("咨询师备注", "remark"),
)

ORDER_COLUMNS = (
    ColumnDef("订单代码", "order_code"),
    ColumnDef("来访电话【必填】", "patient_mobile", True),
    ColumnDef("来访姓名【必填】", "patient_name", True),
    ColumnDef("咨询师电话【必填】", "counselor_mobile", True),
    ColumnDef("咨询师姓名", "counselor_name"),
    ColumnDef("咨询状态【必填】", "order_status", True, ORDER_STATUSES),
    ColumnDef("咨询时间【必填】", "start_time", True),
    ColumnDef("咨询方式", "mode", choices=CONSULTATION_MODES),
    ColumnDef("地点", "location", choices=LOCATIONS),
)

KIND_COLUMNS = {
    "visitors": VISITOR_COLUMNS,
    "counselors": COUNSELOR_COLUMNS,
    "orders": ORDER_COLUMNS,
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _error(sheet: str, cell: str, message: str) -> dict[str, str]:
    return {"sheet": sheet, "cell": cell, "message": message}


def _columns(kind: str) -> tuple[ColumnDef, ...]:
    try:
        return KIND_COLUMNS[kind]
    except KeyError as exc:
        raise ValueError("kind 必须为 visitors、counselors 或 orders") from exc


def _style_sheet(ws, columns: tuple[ColumnDef, ...]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for index, column in enumerate(columns, start=1):
        cell = ws.cell(1, index, column.header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(index)].width = max(16, len(column.header) * 2 + 2)
        if column.choices:
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(column.choices)}"',
                allow_blank=not column.required,
            )
            validation.error = f"请选择：{'、'.join(column.choices)}"
            validation.errorTitle = "值不符合规范"
            validation.showErrorMessage = True
            ws.add_data_validation(validation)
            validation.add(f"{get_column_letter(index)}2:{get_column_letter(index)}1048576")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _workbook_bytes(kind: str, rows: list[list[Any]]) -> bytes:
    columns = _columns(kind)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = {
        "visitors": "来访用户表",
        "counselors": "咨询师用户表",
        "orders": "咨询订单表",
    }[kind]
    _style_sheet(sheet, columns)
    for row in rows:
        sheet.append(row)
    if kind == "orders":
        start_column = next(i for i, col in enumerate(columns, start=1) if col.key == "start_time")
        for cell in sheet[get_column_letter(start_column)][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def template_bytes(kind: str) -> bytes:
    return _workbook_bytes(kind, [])


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = _text(value)
    for pattern in (
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError("须为完整日期和时间，例如 2026-08-12 14:00") from exc


def _parse_price(value: Any) -> Optional[int]:
    text = _text(value).replace(",", "").replace("￥", "").replace("¥", "")
    if not text:
        return None
    try:
        amount = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("须为有效的人民币元金额") from exc
    if amount <= 0:
        raise ValueError("须大于 0")
    cents = int((amount * 100).quantize(Decimal("1")))
    return cents


def _validate_phone(value: Any) -> str:
    mobile = _text(value)
    if not PHONE_RE.fullmatch(mobile):
        raise ValueError("须为 11 位中国大陆手机号")
    return mobile


def _has_role(db: Session, account_id: int, role: str) -> bool:
    return bool(
        db.query(AppRoleBinding)
        .filter(
            AppRoleBinding.AccountId == account_id,
            AppRoleBinding.RoleType == role,
        )
        .first()
    )


def _account_by_mobile(db: Session, mobile: str) -> Optional[AppAccount]:
    return db.query(AppAccount).filter(AppAccount.Mobile == mobile).first()


def _field_cell(columns: tuple[ColumnDef, ...], key: str, row_number: int) -> str:
    index = next(i for i, column in enumerate(columns, start=1) if column.key == key)
    return f"{get_column_letter(index)}{row_number}"


def _validate_row(
    kind: str,
    row: dict[str, Any],
    sheet: str,
    row_number: int,
    db: Session,
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    columns = _columns(kind)
    errors: list[dict[str, str]] = []
    clean = {column.key: _text(row.get(column.key)) for column in columns}

    for column in columns:
        cell = _field_cell(columns, column.key, row_number)
        value = clean[column.key]
        if column.required and not value:
            errors.append(_error(sheet, cell, "不能为空"))
        elif value and column.choices and value not in column.choices:
            errors.append(_error(sheet, cell, f"仅支持：{'、'.join(column.choices)}"))

    phone_keys = {
        "visitors": ("mobile", "bound_counselor_mobile"),
        "counselors": ("mobile",),
        "orders": ("patient_mobile", "counselor_mobile"),
    }[kind]
    for key in phone_keys:
        if not clean[key]:
            continue
        try:
            clean[key] = _validate_phone(clean[key])
        except ValueError as exc:
            errors.append(_error(sheet, _field_cell(columns, key, row_number), str(exc)))

    if kind == "visitors":
        source = clean["patient_source"]
        if source in SOURCE_BY_LABEL:
            clean["patient_source"] = SOURCE_BY_LABEL[source]
        bound_mobile = clean["bound_counselor_mobile"]
        if clean["bound_counselor_name"] and not bound_mobile:
            errors.append(
                _error(
                    sheet,
                    _field_cell(columns, "bound_counselor_mobile", row_number),
                    "填写绑定咨询师名字时必须填写咨询师电话",
                )
            )
        if bound_mobile and PHONE_RE.fullmatch(bound_mobile):
            counselor = _account_by_mobile(db, bound_mobile)
            if not counselor or not _has_role(db, counselor.Id, "Counselor"):
                errors.append(
                    _error(
                        sheet,
                        _field_cell(columns, "bound_counselor_mobile", row_number),
                        "未找到该手机号对应的咨询师",
                    )
                )
            else:
                clean["bound_counselor_id"] = counselor.Id
        clean["is_contract_signed"] = (
            True if clean["contract_status"] == "已签约"
            else False if clean["contract_status"] == "未签约"
            else None
        )

    elif kind == "counselors":
        counselor_type = clean["counselor_type"]
        clean["counselor_type"] = COUNSELOR_TYPE_BY_LABEL.get(counselor_type, counselor_type or None)
        try:
            clean["billing_cents"] = _parse_price(row.get("billing"))
        except ValueError as exc:
            errors.append(_error(sheet, _field_cell(columns, "billing", row_number), str(exc)))

    else:
        status = clean["order_status"] or ORDER_STATUSES[0]
        mode = clean["mode"] or CONSULTATION_MODES[0]
        location = clean["location"] or ("线上" if mode == "视频" else "杨浦咨询中心")
        clean["order_status"] = status
        clean["mode"] = mode
        clean["location"] = location
        if mode == "视频" and location != "线上":
            errors.append(
                _error(sheet, _field_cell(columns, "location", row_number), "视频咨询的地点必须为线上")
            )
        if mode == "线下" and location == "线上":
            errors.append(
                _error(sheet, _field_cell(columns, "location", row_number), "线下咨询不能选择线上地点")
            )
        try:
            clean["start_at"] = _parse_datetime(row.get("start_time"))
            clean["end_at"] = clean["start_at"] + timedelta(minutes=60)
        except ValueError as exc:
            if clean["start_time"]:
                errors.append(_error(sheet, _field_cell(columns, "start_time", row_number), str(exc)))
        for key, role, label in (
            ("patient_mobile", "Patient", "来访"),
            ("counselor_mobile", "Counselor", "咨询师"),
        ):
            mobile = clean[key]
            if mobile and PHONE_RE.fullmatch(mobile):
                account = _account_by_mobile(db, mobile)
                if not account or not _has_role(db, account.Id, role):
                    errors.append(
                        _error(
                            sheet,
                            _field_cell(columns, key, row_number),
                            f"未找到该手机号对应的{label}",
                        )
                    )
                else:
                    clean[f"{key}_id"] = account.Id

    return clean, errors


def _validate_order_business(
    clean: dict[str, Any],
    sheet: str,
    row_number: int,
    db: Session,
) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    columns = ORDER_COLUMNS
    required = ("patient_mobile_id", "counselor_mobile_id", "start_at", "end_at")
    if not all(key in clean for key in required):
        return errors

    code = clean["order_code"] or _generated_order_code(clean)
    clean["order_code"] = code
    if len(code) > 64:
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "order_code", row_number),
                "订单代码不能超过 64 个字符",
            )
        )
    elif db.query(AppOrder).filter(AppOrder.OutTradeNo == code).first():
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "order_code", row_number),
                "订单代码已存在",
            )
        )

    profile = (
        db.query(AppCounselorProfile)
        .filter(AppCounselorProfile.AccountId == clean["counselor_mobile_id"])
        .first()
    )
    if not profile or not profile.Billing:
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "counselor_mobile", row_number),
                "该咨询师缺少有效的基础价格档案",
            )
        )

    duplicate = (
        db.query(AppConsultation)
        .filter(
            AppConsultation.PatientId == clean["patient_mobile_id"],
            AppConsultation.CounselorId == clean["counselor_mobile_id"],
            AppConsultation.StartTime == clean["start_at"],
        )
        .first()
    )
    if duplicate:
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "start_time", row_number),
                "相同来访、咨询师和咨询时间的订单已存在",
            )
        )

    conflict = (
        db.query(AppConsultation)
        .filter(
            AppConsultation.CounselorId == clean["counselor_mobile_id"],
            AppConsultation.Status.in_(["PENDING", "CONFIRMED", "ONGOING", "DONE"]),
            AppConsultation.StartTime < clean["end_at"],
            AppConsultation.EndTime > clean["start_at"],
        )
        .first()
    )
    if conflict and not duplicate:
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "start_time", row_number),
                "该咨询师在此时间段已有订单",
            )
        )
    schedule_conflict = (
        db.query(AppSchedule)
        .filter(
            AppSchedule.CounselorId == clean["counselor_mobile_id"],
            AppSchedule.Status == "BOOKED",
            AppSchedule.StartTime < clean["end_at"],
            AppSchedule.EndTime > clean["start_at"],
        )
        .first()
    )
    if schedule_conflict:
        errors.append(
            _error(
                sheet,
                _field_cell(columns, "start_time", row_number),
                "该咨询师在此时间段已有排期",
            )
        )
    return errors


def parse_and_validate(
    kind: str,
    file_bytes: bytes,
    db: Session,
) -> tuple[list[dict[str, Any]], list[dict[str, str]], int]:
    columns = _columns(kind)
    errors: list[dict[str, str]] = []
    parsed_rows: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return [], [_error("文件", "", f"无法读取 Excel 文件：{exc}")], 0

    for sheet in workbook.worksheets:
        expected_headers = [column.header for column in columns]
        actual_headers = [_text(sheet.cell(1, index).value) for index in range(1, sheet.max_column + 1)]
        for index, expected in enumerate(expected_headers, start=1):
            actual = actual_headers[index - 1] if index <= len(actual_headers) else ""
            if actual != expected:
                errors.append(
                    _error(sheet.title, f"{get_column_letter(index)}1", f"表头应为“{expected}”")
                )
        for index in range(len(expected_headers) + 1, len(actual_headers) + 1):
            if actual_headers[index - 1]:
                errors.append(
                    _error(sheet.title, f"{get_column_letter(index)}1", "存在模板定义之外的列")
                )

        for row_number in range(2, sheet.max_row + 1):
            values = [sheet.cell(row_number, index).value for index in range(1, len(columns) + 1)]
            if not any(_text(value) for value in values):
                continue
            raw = {column.key: values[index] for index, column in enumerate(columns)}
            clean, row_errors = _validate_row(kind, raw, sheet.title, row_number, db)
            errors.extend(row_errors)
            parsed_rows.append({"sheet": sheet.title, "row": row_number, "values": clean})

    return parsed_rows, errors, len(parsed_rows)


def _new_account(
    db: Session,
    mobile: str,
    role: str,
    *,
    real_name: str = "",
    nickname: str = "",
) -> AppAccount:
    digest = hashlib.sha1(f"{role}:{mobile}".encode("utf-8")).hexdigest()[:24]
    account = AppAccount(
        OpenId=f"import-{role.lower()}-{digest}",
        Mobile=mobile,
        ActiveRole=role,
        IsActive=True,
        RealName=real_name or None,
        Nickname=nickname or None,
    )
    db.add(account)
    db.flush()
    db.add(AppRoleBinding(AccountId=account.Id, RoleType=role))
    return account


def _apply_visitor(value: dict[str, Any], db: Session, actor_id: int) -> None:
    account = _new_account(
        db,
        value["mobile"],
        "Patient",
        real_name=value["real_name"],
        nickname=value["nickname"],
    )
    account.PatientSource = value["patient_source"]
    account.PatientSourceDetail = value["patient_source_detail"] or None
    if value["contract_status"]:
        account.IsContractSigned = value["is_contract_signed"]
    if value.get("bound_counselor_id"):
        account.BoundCounselorId = value["bound_counselor_id"]
        account.BoundCounselorChangedAt = datetime.utcnow()
    set_staff_remark(db, account.Id, value["remark"], actor_id)


def _apply_counselor(value: dict[str, Any], db: Session, actor_id: int) -> None:
    account = _new_account(
        db,
        value["mobile"],
        "Counselor",
        real_name=value["name"],
        nickname=value["name"],
    )
    account.Gender = value["gender"] or None
    counselor_type = value["counselor_type"] or "PROFESSIONAL"
    db.add(
        AppCounselorProfile(
            AccountId=account.Id,
            Name=value["name"] or account.RealName or account.Nickname,
            AvatarUrl=DEFAULT_COUNSELOR_PUBLIC_AVATAR,
            CounselorType=counselor_type,
            Billing=(
                value["billing_cents"]
                or default_base_price_cents_for_type(counselor_type)
            ),
            IsActive=True,
        )
    )
    set_staff_remark(db, account.Id, value["remark"], actor_id)


def _generated_order_code(value: dict[str, Any]) -> str:
    identity = (
        f"{value['patient_mobile']}|{value['counselor_mobile']}|"
        f"{value['start_at'].isoformat()}"
    )
    return f"IMPORT-{hashlib.sha1(identity.encode('utf-8')).hexdigest()[:24].upper()}"


def _apply_order(value: dict[str, Any], db: Session) -> None:
    patient = db.query(AppAccount).filter(AppAccount.Id == value["patient_mobile_id"]).one()
    counselor = db.query(AppAccount).filter(AppAccount.Id == value["counselor_mobile_id"]).one()
    profile = (
        db.query(AppCounselorProfile)
        .filter(AppCounselorProfile.AccountId == counselor.Id)
        .one()
    )
    center_id = LOCATION_TO_CENTER[value["location"]]
    note = schedule_note(center_id)
    is_cancelled = value["order_status"] in ORDER_STATUSES[2:]
    consultation_status = (
        "CONFIRMED"
        if value["order_status"] == ORDER_STATUSES[0]
        else "CANCELLED"
        if is_cancelled
        else "DONE"
    )
    schedule = AppSchedule(
        CounselorId=counselor.Id,
        StartTime=value["start_at"],
        EndTime=value["end_at"],
        Status="CANCELLED" if is_cancelled else "BOOKED",
        Note=note,
    )
    db.add(schedule)
    db.flush()
    order = AppOrder(
        AccountId=patient.Id,
        SlotId=schedule.Id,
        OutTradeNo=value["order_code"],
        TotalFee=int(profile.Billing),
        Status=(
            "REFUNDED"
            if value["order_status"] == ORDER_STATUSES[2]
            else "CANCELLED"
            if value["order_status"] == ORDER_STATUSES[3]
            else "PAID"
        ),
        CreatedAt=value["start_at"],
        PaidAt=value["start_at"],
        Description=(
            f"{value['patient_name']}"
            f" / {value['counselor_name'] or profile.Name or counselor.RealName or counselor.Mobile}"
        )[:200],
    )
    db.add(order)
    db.flush()
    consultation = AppConsultation(
        OrderId=order.Id,
        PatientId=patient.Id,
        CounselorId=counselor.Id,
        ScheduleId=schedule.Id,
        Status=consultation_status,
        StartTime=value["start_at"],
        EndTime=value["end_at"],
        Note=note,
    )
    db.add(consultation)


def _duplicate_errors(
    kind: str,
    value: dict[str, Any],
    sheet: str,
    row_number: int,
    db: Session,
) -> list[dict[str, str]]:
    if kind in ("visitors", "counselors"):
        mobile = value.get("mobile")
        if mobile and PHONE_RE.fullmatch(mobile) and _account_by_mobile(db, mobile):
            return [
                _error(
                    sheet,
                    _field_cell(_columns(kind), "mobile", row_number),
                    "该手机号已存在，导入仅支持新增",
                )
            ]
        return []
    return _validate_order_business(value, sheet, row_number, db)


def _result_row(
    sheet: str,
    row_number: int,
    status: str,
    errors: list[dict[str, str]],
) -> dict[str, Any]:
    message = (
        "导入成功"
        if status == "IMPORTED"
        else "；".join(error["message"] for error in errors)
    )
    return {
        "sheet": sheet,
        "row": row_number,
        "rowNumber": row_number,
        "status": status,
        "message": message,
        "errors": errors,
    }


def import_workbook(
    kind: str,
    file_bytes: bytes,
    db: Session,
    actor_id: int,
) -> dict[str, Any]:
    columns = _columns(kind)
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        errors = [_error("文件", "", f"无法读取 Excel 文件：{exc}")]
        return {
            "message": "导入失败：无法读取 Excel 文件",
            "totalRows": 0,
            "importedCount": 0,
            "rejectedCount": 0,
            "failedCount": 0,
            "errors": errors,
            "rows": [],
        }

    source_rows: list[dict[str, Any]] = []
    workbook_errors: list[dict[str, str]] = []
    expected_headers = [column.header for column in columns]
    for sheet in workbook.worksheets:
        header_errors: list[dict[str, str]] = []
        actual_headers = [
            _text(sheet.cell(1, index).value)
            for index in range(1, sheet.max_column + 1)
        ]
        for index, expected in enumerate(expected_headers, start=1):
            actual = actual_headers[index - 1] if index <= len(actual_headers) else ""
            if actual != expected:
                header_errors.append(
                    _error(sheet.title, f"{get_column_letter(index)}1", f"表头应为“{expected}”")
                )
        for index in range(len(expected_headers) + 1, len(actual_headers) + 1):
            if actual_headers[index - 1]:
                header_errors.append(
                    _error(sheet.title, f"{get_column_letter(index)}1", "存在模板定义之外的列")
                )
        workbook_errors.extend(header_errors)
        for row_number in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row_number, index).value
                for index in range(1, len(columns) + 1)
            ]
            if not any(_text(value) for value in values):
                continue
            source_rows.append(
                {
                    "sheet": sheet.title,
                    "row": row_number,
                    "raw": {
                        column.key: values[index]
                        for index, column in enumerate(columns)
                    },
                    "header_errors": list(header_errors),
                }
            )

    result_rows: list[dict[str, Any]] = []
    all_errors: list[dict[str, str]] = []
    successful_keys: set[tuple[Any, ...]] = set()
    if not source_rows:
        all_errors.extend(workbook_errors)
        all_errors.append(_error("文件", "", "导入文件没有数据行"))

    for item in source_rows:
        sheet = item["sheet"]
        row_number = item["row"]
        clean, row_errors = _validate_row(
            kind, item["raw"], sheet, row_number, db
        )
        row_errors = list(item["header_errors"]) + row_errors
        key: tuple[Any, ...]
        if kind in ("visitors", "counselors"):
            key = (clean.get("mobile"),)
        else:
            key = (
                clean.get("patient_mobile"),
                clean.get("counselor_mobile"),
                clean.get("start_at"),
            )
        if all(key) and key in successful_keys:
            duplicate_key = "mobile" if kind != "orders" else "start_time"
            row_errors.append(
                _error(
                    sheet,
                    _field_cell(columns, duplicate_key, row_number),
                    "同一文件中与此前成功导入的行重复",
                )
            )
        row_errors.extend(_duplicate_errors(kind, clean, sheet, row_number, db))
        if row_errors:
            db.rollback()
            all_errors.extend(row_errors)
            result_rows.append(_result_row(sheet, row_number, "REJECTED", row_errors))
            continue

        # 在真正写入前再次查询，缩短校验与写入之间的竞态窗口。
        second_errors = _duplicate_errors(kind, clean, sheet, row_number, db)
        if second_errors:
            db.rollback()
            all_errors.extend(second_errors)
            result_rows.append(_result_row(sheet, row_number, "REJECTED", second_errors))
            continue
        try:
            if kind == "visitors":
                _apply_visitor(clean, db, actor_id)
            elif kind == "counselors":
                _apply_counselor(clean, db, actor_id)
            else:
                _apply_order(clean, db)
            db.commit()
        except Exception as exc:
            db.rollback()
            failed_errors = [_error(sheet, "", str(exc))]
            all_errors.extend(failed_errors)
            result_rows.append(_result_row(sheet, row_number, "FAILED", failed_errors))
            continue
        successful_keys.add(key)
        result_rows.append(_result_row(sheet, row_number, "IMPORTED", []))

    imported = sum(row["status"] == "IMPORTED" for row in result_rows)
    rejected = sum(row["status"] == "REJECTED" for row in result_rows)
    failed = sum(row["status"] == "FAILED" for row in result_rows)
    return {
        "message": f"导入完成：成功 {imported} 行，拒绝 {rejected} 行，失败 {failed} 行",
        "totalRows": len(source_rows),
        "importedCount": imported,
        "rejectedCount": rejected,
        "failedCount": failed,
        "errors": all_errors,
        "rows": result_rows,
    }


def _role_accounts(db: Session, role: str) -> list[AppAccount]:
    return (
        db.query(AppAccount)
        .join(AppRoleBinding, AppRoleBinding.AccountId == AppAccount.Id)
        .filter(AppRoleBinding.RoleType == role)
        .distinct()
        .order_by(AppAccount.Id.asc())
        .all()
    )


def _export_visitors(db: Session) -> list[list[Any]]:
    accounts = _role_accounts(db, "Patient")
    remarks = get_staff_remarks_map(db, [account.Id for account in accounts])
    bound_ids = {account.BoundCounselorId for account in accounts if account.BoundCounselorId}
    bound = {
        account.Id: account
        for account in db.query(AppAccount).filter(AppAccount.Id.in_(bound_ids)).all()
    } if bound_ids else {}
    rows = []
    for account in accounts:
        counselor = bound.get(account.BoundCounselorId)
        rows.append(
            [
                account.Mobile or "",
                account.RealName or "",
                account.Nickname or "",
                "已签约" if account.IsContractSigned else "未签约",
                (
                    counselor.RealName or counselor.Nickname or ""
                    if counselor
                    else ""
                ),
                counselor.Mobile or "" if counselor else "",
                PATIENT_SOURCES.get(
                    normalize_patient_source(account.PatientSource),
                    account.PatientSource or "",
                ),
                account.PatientSourceDetail or "",
                remarks.get(account.Id, ""),
            ]
        )
    return rows


def _export_counselors(db: Session) -> list[list[Any]]:
    accounts = _role_accounts(db, "Counselor")
    remarks = get_staff_remarks_map(db, [account.Id for account in accounts])
    profiles = {
        profile.AccountId: profile
        for profile in db.query(AppCounselorProfile)
        .filter(AppCounselorProfile.AccountId.in_([account.Id for account in accounts]))
        .all()
    } if accounts else {}
    rows = []
    for account in accounts:
        profile = profiles.get(account.Id)
        rows.append(
            [
                account.Mobile or "",
                (profile.Name if profile else None) or account.RealName or account.Nickname or "",
                account.Gender or "",
                COUNSELOR_TYPES.get(
                    profile.CounselorType if profile else None,
                    profile.CounselorType if profile else "",
                ),
                Decimal(profile.Billing) / Decimal(100) if profile and profile.Billing is not None else "",
                remarks.get(account.Id, ""),
            ]
        )
    return rows


def _export_orders(
    db: Session,
    start_date: Optional[date],
    end_date: Optional[date],
) -> list[list[Any]]:
    query = db.query(AppConsultation).filter(
        AppConsultation.Status.in_(["PENDING", "CONFIRMED", "ONGOING", "DONE", "CANCELLED", "CANCELED"]),
        AppConsultation.StartTime.isnot(None),
    )
    if start_date:
        query = query.filter(
            AppConsultation.StartTime >= datetime.combine(start_date, time.min)
        )
    if end_date:
        query = query.filter(
            AppConsultation.StartTime <= datetime.combine(end_date, time.max)
        )
    consultations = query.order_by(AppConsultation.StartTime.asc(), AppConsultation.Id.asc()).all()
    account_ids = {
        account_id
        for consultation in consultations
        for account_id in (consultation.PatientId, consultation.CounselorId)
    }
    accounts = {
        account.Id: account
        for account in db.query(AppAccount).filter(AppAccount.Id.in_(account_ids)).all()
    } if account_ids else {}
    order_ids = {consultation.OrderId for consultation in consultations if consultation.OrderId}
    orders = {
        order.Id: order
        for order in db.query(AppOrder).filter(AppOrder.Id.in_(order_ids)).all()
    } if order_ids else {}
    profiles = {
        profile.AccountId: profile
        for profile in db.query(AppCounselorProfile)
        .filter(
            AppCounselorProfile.AccountId.in_(
                [consultation.CounselorId for consultation in consultations]
            )
        )
        .all()
    } if consultations else {}

    rows = []
    for consultation in consultations:
        patient = accounts.get(consultation.PatientId)
        counselor = accounts.get(consultation.CounselorId)
        profile = profiles.get(consultation.CounselorId)
        order = orders.get(consultation.OrderId)
        center_id = parse_center_id(consultation.Note)
        location = CENTER_TO_LOCATION.get(center_id, "线上" if center_id == "video" else "")
        status = (
            ORDER_STATUSES[2]
            if order and order.Status == "REFUNDED"
            else ORDER_STATUSES[3]
            if order and order.Status == "CANCELLED"
            else ORDER_STATUSES[3]
            if consultation.Status in ("CANCELLED", "CANCELED")
            else ORDER_STATUSES[1]
            if consultation.Status == "DONE"
            else ORDER_STATUSES[0]
        )
        rows.append(
            [
                order.OutTradeNo if order else "",
                patient.Mobile if patient else "",
                (patient.RealName or patient.Nickname or "") if patient else "",
                counselor.Mobile if counselor else "",
                (
                    (profile.Name if profile else None)
                    or (counselor.RealName if counselor else None)
                    or (counselor.Nickname if counselor else None)
                    or ""
                ),
                status,
                consultation.StartTime,
                "视频" if center_id == "video" else "线下",
                location,
            ]
        )
    return rows


def export_bytes(
    kind: str,
    db: Session,
    *,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> bytes:
    if kind == "visitors":
        rows = _export_visitors(db)
    elif kind == "counselors":
        rows = _export_counselors(db)
    elif kind == "orders":
        if start_date and end_date and start_date > end_date:
            raise ValueError("startDate 不能晚于 endDate")
        rows = _export_orders(db, start_date, end_date)
    else:
        _columns(kind)
        raise AssertionError("unreachable")
    return _workbook_bytes(kind, rows)
