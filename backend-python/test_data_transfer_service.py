import unittest
from datetime import date, datetime
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from data_transfer_service import (
    KIND_COLUMNS,
    ORDER_STATUSES,
    export_bytes,
    import_workbook,
    template_bytes,
)
from database import Base
from models import (
    AppAccount,
    AppCaseRecord,
    AppConsultation,
    AppCounselorProfile,
    AppOrder,
    AppRoleBinding,
    AppSchedule,
    AppStaffAccountRemark,
)


TABLES = [
    AppAccount.__table__,
    AppRoleBinding.__table__,
    AppCounselorProfile.__table__,
    AppStaffAccountRemark.__table__,
    AppOrder.__table__,
    AppSchedule.__table__,
    AppConsultation.__table__,
    AppCaseRecord.__table__,
]


def workbook_bytes(kind, rows, *, second_sheet_rows=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据一"
    headers = [column.header for column in KIND_COLUMNS[kind]]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    if second_sheet_rows is not None:
        second = workbook.create_sheet("数据二")
        second.append(headers)
        for row in second_sheet_rows:
            second.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class DataTransferServiceTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine, tables=TABLES)
        self.Session = sessionmaker(bind=self.engine, autoflush=False)
        self.db = self.Session()
        self.db.add(AppAccount(Id=99, Mobile="13800000099", ActiveRole="Admin"))
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def add_role_account(self, account_id, mobile, role, name):
        account = AppAccount(
            Id=account_id,
            Mobile=mobile,
            RealName=name,
            Nickname=name,
            ActiveRole=role,
            IsActive=True,
        )
        self.db.add(account)
        self.db.add(AppRoleBinding(AccountId=account_id, RoleType=role))
        self.db.flush()
        return account

    def test_templates_use_the_shared_column_definitions(self):
        sheet_names = {
            "visitors": "来访用户表",
            "counselors": "咨询师用户表",
            "orders": "咨询订单表",
        }
        for kind, columns in KIND_COLUMNS.items():
            workbook = load_workbook(BytesIO(template_bytes(kind)))
            headers = [cell.value for cell in workbook.active[1]]
            self.assertEqual(headers, [column.header for column in columns])
            self.assertEqual(workbook.active.title, sheet_names[kind])
        self.assertIn("来访类别【必填】", [column.header for column in KIND_COLUMNS["visitors"]])
        self.assertIn("来访来源", [column.header for column in KIND_COLUMNS["visitors"]])
        self.assertIn("咨询师姓名【必填】", [column.header for column in KIND_COLUMNS["counselors"]])
        self.assertIn("咨询师性别", [column.header for column in KIND_COLUMNS["counselors"]])
        self.assertIn("来访姓名【必填】", [column.header for column in KIND_COLUMNS["orders"]])
        self.assertIn("咨询状态【必填】", [column.header for column in KIND_COLUMNS["orders"]])
        self.assertTrue(
            next(column for column in KIND_COLUMNS["orders"] if column.key == "order_status").required
        )
        self.assertNotIn("主观记录", [column.header for column in KIND_COLUMNS["orders"]])

    def test_validation_collects_errors_from_every_sheet_without_writes(self):
        content = workbook_bytes(
            "visitors",
            [["123", "", "", "", "", "", "未知来源", "", ""]],
            second_sheet_rows=[["", "", "", "", "", "", "", "", "仅备注"]],
        )
        result = import_workbook("visitors", content, self.db, 99)

        self.assertEqual(result["importedCount"], 0)
        self.assertEqual(result["totalRows"], 2)
        self.assertGreaterEqual(len(result["errors"]), 4)
        self.assertEqual(
            self.db.query(AppRoleBinding).filter(AppRoleBinding.RoleType == "Patient").count(),
            0,
        )

    def test_visitor_import_rejects_non_whitelisted_source_detail(self):
        result = import_workbook(
            "visitors",
            workbook_bytes(
                "visitors",
                [["13800000001", "来访甲", "", "", "", "", "正价", "医院转介", ""]],
            ),
            self.db,
            99,
        )
        self.assertEqual(result["importedCount"], 0)
        self.assertTrue(any("仅支持" in error["message"] for error in result["errors"]))

    def test_visitor_and_counselor_import_create_profiles_prices_and_remarks(self):
        visitor_result = import_workbook(
            "visitors",
            workbook_bytes(
                "visitors",
                [[
                    "13800000001",
                    "来访甲",
                    "小甲",
                    "已签约",
                    "",
                    "",
                    "正价",
                    "医院转出",
                    "来访备注",
                ]],
            ),
            self.db,
            99,
        )
        counselor_result = import_workbook(
            "counselors",
            workbook_bytes(
                "counselors",
                [["13800000002", "咨询师甲", "女", "专业咨询师", 688.5, "咨询师备注"]],
            ),
            self.db,
            99,
        )

        self.assertEqual(visitor_result["importedCount"], 1)
        self.assertEqual(counselor_result["importedCount"], 1)
        visitor = self.db.query(AppAccount).filter(AppAccount.Mobile == "13800000001").one()
        counselor = self.db.query(AppAccount).filter(AppAccount.Mobile == "13800000002").one()
        profile = (
            self.db.query(AppCounselorProfile)
            .filter(AppCounselorProfile.AccountId == counselor.Id)
            .one()
        )
        self.assertTrue(visitor.IsContractSigned)
        self.assertEqual(visitor.PatientSource, "PROFESSIONAL")
        self.assertEqual(visitor.PatientSourceDetail, "医院转出")
        self.assertEqual(counselor.Gender, "女")
        self.assertEqual(profile.Billing, 68850)
        self.assertEqual(
            self.db.query(AppStaffAccountRemark)
            .filter(AppStaffAccountRemark.AccountId == visitor.Id)
            .one()
            .Remark,
            "来访备注",
        )

    def test_order_import_creates_single_sixty_minute_transaction_graph(self):
        patient = self.add_role_account(1, "13800000001", "Patient", "来访甲")
        counselor = self.add_role_account(2, "13800000002", "Counselor", "咨询师甲")
        self.db.add(
            AppCounselorProfile(
                AccountId=counselor.Id,
                Name="咨询师甲",
                CounselorType="PROFESSIONAL",
                Billing=70000,
                IsActive=True,
            )
        )
        self.db.commit()

        result = import_workbook(
            "orders",
            workbook_bytes(
                "orders",
                [[
                    "ORDER-001",
                    patient.Mobile,
                    "小甲",
                    counselor.Mobile,
                    "咨询师甲",
                    ORDER_STATUSES[1],
                    datetime(2026, 8, 12, 14, 0),
                    "视频",
                    "线上",
                ]],
            ),
            self.db,
            99,
        )

        self.assertEqual(result["errors"], [])
        self.assertEqual(result["importedCount"], 1)
        order = self.db.query(AppOrder).one()
        schedule = self.db.query(AppSchedule).one()
        consultation = self.db.query(AppConsultation).one()
        self.assertEqual(order.TotalFee, 70000)
        self.assertEqual(schedule.EndTime - schedule.StartTime, consultation.EndTime - consultation.StartTime)
        self.assertEqual(consultation.EndTime - consultation.StartTime, datetime(2026, 8, 12, 15, 0) - datetime(2026, 8, 12, 14, 0))
        self.assertEqual(consultation.Status, "DONE")
        self.assertEqual(self.db.query(AppCaseRecord).count(), 0)

    def test_order_without_code_rejects_duplicate_and_imports_refund_status(self):
        patient = self.add_role_account(1, "13800000001", "Patient", "来访甲")
        counselor = self.add_role_account(2, "13800000002", "Counselor", "咨询师甲")
        self.db.add(
            AppCounselorProfile(
                AccountId=counselor.Id,
                Name="咨询师甲",
                CounselorType="PROFESSIONAL",
                Billing=70000,
                IsActive=True,
            )
        )
        self.db.commit()
        base_row = [
            "",
            patient.Mobile,
            "小甲",
            counselor.Mobile,
            "咨询师甲",
            ORDER_STATUSES[0],
            datetime(2026, 8, 12, 14, 0),
            "视频",
            "线上",
        ]

        first = import_workbook(
            "orders", workbook_bytes("orders", [base_row]), self.db, 99
        )
        second = import_workbook(
            "orders", workbook_bytes("orders", [base_row]), self.db, 99
        )

        self.assertEqual(first["importedCount"], 1)
        self.assertEqual(second["importedCount"], 0)
        self.assertEqual(second["rejectedCount"], 1)
        self.assertEqual(second["rows"][0]["status"], "REJECTED")
        self.assertEqual(self.db.query(AppOrder).count(), 1)
        self.assertEqual(self.db.query(AppConsultation).count(), 1)
        self.assertEqual(self.db.query(AppSchedule).count(), 1)

        refunded_row = list(base_row)
        refunded_row[5] = ORDER_STATUSES[2]
        refunded_row[6] = datetime(2026, 8, 13, 14, 0)
        refunded = import_workbook(
            "orders", workbook_bytes("orders", [refunded_row]), self.db, 99
        )
        self.assertEqual(refunded["importedCount"], 1)
        imported_refund = self.db.query(AppOrder).order_by(AppOrder.Id.desc()).first()
        self.assertEqual(imported_refund.Status, "REFUNDED")
        cancelled_row = list(base_row)
        cancelled_row[5] = ORDER_STATUSES[3]
        cancelled_row[6] = datetime(2026, 8, 14, 14, 0)
        cancelled = import_workbook(
            "orders", workbook_bytes("orders", [cancelled_row]), self.db, 99
        )
        self.assertEqual(cancelled["importedCount"], 1)
        imported_cancel = self.db.query(AppOrder).order_by(AppOrder.Id.desc()).first()
        self.assertEqual(imported_cancel.Status, "CANCELLED")
        self.assertEqual(self.db.query(AppOrder).count(), 3)

    def test_visitor_partial_success_duplicate_and_statistics(self):
        self.db.add(
            AppAccount(
                Id=10,
                Mobile="13800000010",
                ActiveRole="Patient",
                IsActive=False,
                DeletedAt=datetime(2026, 1, 1),
            )
        )
        self.db.commit()
        content = workbook_bytes(
            "visitors",
            [
                ["13800000001", "甲", "", "", "", "", "正价", "", ""],
                ["13800000001", "重复甲", "", "", "", "", "正价", "", ""],
                ["13800000010", "已删除账号", "", "", "", "", "正价", "", ""],
                ["13800000003", "丙", "", "", "", "", "未知来源", "", ""],
                ["13800000004", "丁", "", "", "", "", "正价", "", ""],
            ],
        )

        result = import_workbook("visitors", content, self.db, 99)

        self.assertEqual(result["totalRows"], 5)
        self.assertEqual(result["importedCount"], 2)
        self.assertEqual(result["rejectedCount"], 3)
        self.assertEqual(result["failedCount"], 0)
        self.assertEqual(
            [row["status"] for row in result["rows"]],
            ["IMPORTED", "REJECTED", "REJECTED", "REJECTED", "IMPORTED"],
        )
        self.assertEqual(
            self.db.query(AppAccount)
            .filter(AppAccount.Mobile.in_(["13800000001", "13800000004"]))
            .count(),
            2,
        )

    def test_row_failure_rolls_back_only_current_row(self):
        content = workbook_bytes(
            "visitors",
            [
                ["13800000001", "甲", "", "", "", "", "正价", "", ""],
                ["13800000002", "乙", "", "", "", "", "正价", "", ""],
                ["13800000003", "丙", "", "", "", "", "正价", "", ""],
            ],
        )
        from data_transfer_service import _apply_visitor

        def fail_middle(value, db, actor_id):
            if value["mobile"] == "13800000002":
                raise RuntimeError("模拟当前行写入失败")
            return _apply_visitor(value, db, actor_id)

        with patch("data_transfer_service._apply_visitor", side_effect=fail_middle):
            result = import_workbook("visitors", content, self.db, 99)

        self.assertEqual(result["importedCount"], 2)
        self.assertEqual(result["rejectedCount"], 0)
        self.assertEqual(result["failedCount"], 1)
        self.assertEqual(
            [row["status"] for row in result["rows"]],
            ["IMPORTED", "FAILED", "IMPORTED"],
        )
        self.assertIsNotNone(
            self.db.query(AppAccount).filter(AppAccount.Mobile == "13800000001").first()
        )
        self.assertIsNone(
            self.db.query(AppAccount).filter(AppAccount.Mobile == "13800000002").first()
        )
        self.assertIsNotNone(
            self.db.query(AppAccount).filter(AppAccount.Mobile == "13800000003").first()
        )

    def test_order_duplicate_checks_include_cancelled_consultations(self):
        patient = self.add_role_account(1, "13800000001", "Patient", "来访甲")
        counselor = self.add_role_account(2, "13800000002", "Counselor", "咨询师甲")
        self.db.add(
            AppCounselorProfile(
                AccountId=counselor.Id,
                Name="咨询师甲",
                CounselorType="PROFESSIONAL",
                Billing=70000,
                IsActive=True,
            )
        )
        self.db.add(
            AppConsultation(
                PatientId=patient.Id,
                CounselorId=counselor.Id,
                Status="CANCELLED",
                StartTime=datetime(2026, 8, 12, 14, 0),
                EndTime=datetime(2026, 8, 12, 15, 0),
            )
        )
        self.db.commit()

        result = import_workbook(
            "orders",
            workbook_bytes(
                "orders",
                [[
                    "NEW-CODE",
                    patient.Mobile,
                    "来访甲",
                    counselor.Mobile,
                    "",
                    ORDER_STATUSES[0],
                    datetime(2026, 8, 12, 14, 0),
                    "视频",
                    "线上",
                ]],
            ),
            self.db,
            99,
        )

        self.assertEqual(result["importedCount"], 0)
        self.assertEqual(result["rejectedCount"], 1)
        self.assertTrue(
            any("订单已存在" in error["message"] for error in result["errors"])
        )
        self.assertEqual(self.db.query(AppOrder).count(), 0)

    def test_order_export_filters_inclusive_dates_and_derives_status(self):
        patient = self.add_role_account(1, "13800000001", "Patient", "来访甲")
        counselor = self.add_role_account(2, "13800000002", "Counselor", "咨询师甲")
        self.db.add(
            AppCounselorProfile(
                AccountId=counselor.Id,
                Name="咨询师甲",
                CounselorType="PROFESSIONAL",
                Billing=60000,
                IsActive=True,
            )
        )
        order = AppOrder(
            Id=1,
            AccountId=patient.Id,
            OutTradeNo="ORDER-EXPORT",
            TotalFee=60000,
            Status="PAID",
        )
        self.db.add(order)
        self.db.add(
            AppConsultation(
                Id=1,
                OrderId=1,
                PatientId=patient.Id,
                CounselorId=counselor.Id,
                Status="DONE",
                StartTime=datetime(2026, 8, 12, 23, 59),
                EndTime=datetime(2026, 8, 13, 0, 59),
                Note="center:video",
            )
        )
        self.db.add(
            AppConsultation(
                Id=2,
                PatientId=patient.Id,
                CounselorId=counselor.Id,
                Status="PENDING",
                StartTime=datetime(2026, 8, 12, 10, 0),
                EndTime=datetime(2026, 8, 12, 11, 0),
                Note="center:yangpu",
            )
        )
        self.db.add(
            AppConsultation(
                Id=3,
                PatientId=patient.Id,
                CounselorId=counselor.Id,
                Status="CANCELLED",
                StartTime=datetime(2026, 8, 12, 12, 0),
                EndTime=datetime(2026, 8, 12, 13, 0),
                Note="center:yangpu",
            )
        )
        self.db.commit()

        workbook = load_workbook(
            BytesIO(
                export_bytes(
                    "orders",
                    self.db,
                    start_date=date(2026, 8, 12),
                    end_date=date(2026, 8, 12),
                )
            ),
            data_only=True,
        )
        values = list(workbook.active.values)
        self.assertEqual(len(values), 4)
        self.assertEqual(values[1][5], ORDER_STATUSES[0])
        self.assertEqual(values[2][5], ORDER_STATUSES[3])
        self.assertEqual(values[3][0], "ORDER-EXPORT")
        self.assertEqual(values[3][5], ORDER_STATUSES[1])


if __name__ == "__main__":
    unittest.main()
